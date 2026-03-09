import io
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.budget_item import PaymentRhythm


# ── Colors (web design system, light mode) ────────────────────────────────────
ACCENT_HEX    = "1A1A1A"
INCOME_HEX    = "166534"
INCOME_BG     = "F0FDF4"
INCOME_ROW    = "DCFCE7"
EXPENSE_HEX   = "B91C1C"
EXPENSE_BG    = "FEF2F2"
EXPENSE_ROW   = "FFE4E4"
HEADER_FG     = "FFFFFF"
MUTED_HEX     = "6B7280"
BORDER_HEX    = "B8B4AC"
BG_HEX        = "ECEAE4"
SECTION_BG    = "FAF9F7"
SUMMARY_BG    = "F3F2EF"


RHYTHM_LABELS = {
    PaymentRhythm.MONTHLY:       "Monatlich",
    PaymentRhythm.QUARTERLY:     "Vierteljährlich",
    PaymentRhythm.SEMI_ANNUALLY: "Halbjährlich",
    PaymentRhythm.ANNUALLY:      "Jährlich",
}

RHYTHM_LIST = '"Monatlich,Vierteljährlich,Halbjährlich,Jährlich"'
TYPE_LIST   = '"Einnahme,Ausgabe"'

TYPE_LABELS = {
    "income":  "Einnahme",
    "expense": "Ausgabe",
}


def _rhythm_label(rhythm) -> str:
    if isinstance(rhythm, PaymentRhythm):
        return RHYTHM_LABELS.get(rhythm, str(rhythm))
    return RHYTHM_LABELS.get(PaymentRhythm(rhythm), str(rhythm))


def _type_label(item) -> str:
    raw = str(getattr(item.get("type"), "value", item.get("type")))
    return TYPE_LABELS.get(raw, raw)


def _border(color=BORDER_HEX) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _monthly_formula(row: int) -> str:
    """Monatlicher Betrag = Betrag / Teiler.
    Uses LEFT() on the first ASCII character to avoid umlauts in formula strings.
    M=Monatlich/1, V=Vierteljährlich/3, H=Halbjährlich/6, J=Jährlich/12.
    Returns empty string if Betrag is empty (for blank input rows).
    """
    return (
        f'=IF(D{row}="","",D{row}/'
        f'CHOOSE(MATCH(LEFT(E{row},1),'
        f'{{"M","V","H","J"}},0),1,3,6,12))'
    )


def generate_plan_excel(
    plan_name: str,
    plan_description: str | None,
    items: list[dict],
) -> bytes:
    wb = Workbook()

    _build_items_sheet(wb, items)
    _build_auswertung_sheet(wb, plan_name, plan_description, items)

    # Remove default sheet if still there
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Sheet 1: Budget-Posten ────────────────────────────────────────────────────

def _build_items_sheet(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Budget-Posten")

    # Column widths
    col_widths = {"A": 12, "B": 32, "C": 22, "D": 16, "E": 20, "F": 18, "G": 32}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Header row ───────────────────────────────────────────────────────────
    headers = ["Typ", "Beschreibung", "Kategorie", "Betrag (€)",
               "Zahlungsrhythmus", "Monatl. Betrag (€)", "Bemerkung"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
        cell.fill      = _fill(ACCENT_HEX)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[1].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    income_items  = [i for i in items if str(getattr(i.get("type"), "value", i.get("type"))) == "income"]
    expense_items = [i for i in items if str(getattr(i.get("type"), "value", i.get("type"))) == "expense"]
    sorted_items  = income_items + expense_items

    currency_fmt = '#,##0.00 "€"'

    for row_idx, item in enumerate(sorted_items, 2):
        is_income = str(getattr(item.get("type"), "value", item.get("type"))) == "income"
        row_fill  = _fill(INCOME_ROW if is_income else EXPENSE_ROW)
        row_font  = Font(name="Calibri", size=10,
                         color=INCOME_HEX if is_income else EXPENSE_HEX)
        val_font  = Font(name="Calibri", size=10, bold=True,
                         color=INCOME_HEX if is_income else EXPENSE_HEX)

        values = [
            _type_label(item),
            item["description"],
            item.get("category_name", str(item.get("category_id", ""))),
            float(item["amount"]),
            _rhythm_label(item["payment_rhythm"]),
            None,                    # formula set below
            item.get("note") or "",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = val_font if col_idx in (4, 6) else row_font
            cell.fill      = row_fill
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 7)))
            if col_idx in (4, 6):
                cell.number_format = currency_fmt
            if col_idx in (1, 5):
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Monthly amount formula
        formula_cell = ws.cell(row=row_idx, column=6, value=_monthly_formula(row_idx))
        formula_cell.font         = val_font
        formula_cell.fill         = row_fill
        formula_cell.border       = _border()
        formula_cell.number_format = currency_fmt
        formula_cell.alignment    = Alignment(horizontal="right", vertical="center")

    # ── Empty rows for adding items ───────────────────────────────────────────
    last_data_row = len(sorted_items) + 1
    empty_fill    = _fill(SECTION_BG)
    empty_font    = Font(name="Calibri", size=10, color=MUTED_HEX, italic=True)

    for row_idx in range(last_data_row + 1, last_data_row + 11):
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = empty_fill
            cell.border = _border("D4D0C9")
            cell.font   = empty_font
            if col_idx == 6 and row_idx > 1:
                cell.value         = _monthly_formula(row_idx)
                cell.number_format = currency_fmt
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            if col_idx == 4 and row_idx > 1:
                cell.number_format = currency_fmt

    # ── Data validation dropdowns ─────────────────────────────────────────────
    last_row = last_data_row + 10

    dv_type = DataValidation(type="list", formula1=TYPE_LIST, allow_blank=True,
                              showDropDown=False, showErrorMessage=True,
                              errorTitle="Ungültiger Typ",
                              error="Bitte 'Einnahme' oder 'Ausgabe' wählen.")
    dv_type.sqref = f"A2:A{last_row}"
    ws.add_data_validation(dv_type)

    dv_rhythm = DataValidation(type="list", formula1=RHYTHM_LIST, allow_blank=True,
                                showDropDown=False, showErrorMessage=True,
                                errorTitle="Ungültiger Rhythmus",
                                error="Bitte einen Zahlungsrhythmus wählen.")
    dv_rhythm.sqref = f"E2:E{last_row}"
    ws.add_data_validation(dv_rhythm)


# ── Sheet 2: Auswertung ───────────────────────────────────────────────────────

def _build_auswertung_sheet(
    wb: Workbook,
    plan_name: str,
    plan_description: str | None,
    items: list[dict],
) -> None:
    ws = wb.create_sheet("Auswertung")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12

    income_items  = [i for i in items if str(getattr(i.get("type"), "value", i.get("type"))) == "income"]
    expense_items = [i for i in items if str(getattr(i.get("type"), "value", i.get("type"))) == "expense"]

    # Row ranges in Budget-Posten (row 1 = header, income first, then expense)
    income_first  = 2
    income_last   = 1 + len(income_items)
    expense_first = 2 + len(income_items)
    expense_last  = 1 + len(income_items) + len(expense_items)

    # SUM formulas over fixed row ranges - avoids inlineStr SUMIF issues
    def _sum_range(first: int, last: int) -> str:
        if first > last:
            return "=0"
        return f"=SUM('Budget-Posten'!F{first}:F{last})"

    def _sumif_cat(first: int, last: int, cat: str) -> str:
        if first > last:
            return "=0"
        return (
            f"=SUMIF('Budget-Posten'!C{first}:C{last},"
            f"\"{cat}\",'Budget-Posten'!F{first}:F{last})"
        )

    currency_fmt = '#,##0.00 "€"'
    pct_fmt      = "0.0%"
    row          = 1

    def _write_cell(r, c, value, font, fill_hex, fmt=None, align=None, border=True):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font
        cell.fill = _fill(fill_hex)
        if border:
            cell.border = _border()
        if fmt:
            cell.number_format = fmt
        if align:
            cell.alignment = align
        return cell

    def _section_header(label: str, bg: str):
        nonlocal row
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill   = _fill(bg)
            ws.cell(row=row, column=c).border = _border()
        cell = ws.cell(row=row, column=1, value=label)
        cell.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 20
        row += 1

    def _kv_row(label: str, formula: str, fmt: str, color: str, bold: bool = False):
        nonlocal row
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font      = Font(name="Calibri", size=10, bold=bold, color=MUTED_HEX)
        lbl.fill      = _fill(SUMMARY_BG)
        lbl.border    = _border()
        lbl.alignment = Alignment(vertical="center")

        val = ws.cell(row=row, column=2, value=formula)
        val.font          = Font(name="Calibri", size=10, bold=bold, color=color)
        val.fill          = _fill(SUMMARY_BG)
        val.border        = _border()
        val.number_format = fmt
        val.alignment     = Alignment(horizontal="right", vertical="center")

        ws.cell(row=row, column=3).fill   = _fill(SUMMARY_BG)
        ws.cell(row=row, column=3).border = _border()
        row += 1
        return row - 1  # return the row that was just written

    # ── Plan header ──────────────────────────────────────────────────────────
    title = ws.cell(row=row, column=1, value=plan_name)
    title.font      = Font(name="Calibri", bold=True, size=16, color=ACCENT_HEX)
    title.alignment = Alignment(vertical="center")
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 28
    row += 1

    if plan_description:
        desc = ws.cell(row=row, column=1, value=plan_description)
        desc.font      = Font(name="Calibri", size=10, color=MUTED_HEX)
        desc.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

    dt = ws.cell(row=row, column=1,
                 value=f"Exportiert am {date.today().strftime('%d.%m.%Y')}")
    dt.font      = Font(name="Calibri", size=9, italic=True, color=MUTED_HEX)
    ws.merge_cells(f"A{row}:C{row}")
    row += 2

    # ── Monatliche Übersicht ──────────────────────────────────────────────────
    _section_header("Monatliche Übersicht", ACCENT_HEX)

    income_result_row  = _kv_row("Monatliche Einnahmen",
                                  _sum_range(income_first, income_last),
                                  currency_fmt, INCOME_HEX, True)
    expense_result_row = _kv_row("Monatliche Ausgaben",
                                  _sum_range(expense_first, expense_last),
                                  currency_fmt, EXPENSE_HEX, True)
    _kv_row("Monatliche Bilanz",
            f"=B{income_result_row}-B{expense_result_row}",
            currency_fmt, ACCENT_HEX, True)
    row += 1  # spacer

    # ── Subheader helper ──────────────────────────────────────────────────────
    def _cat_subheader(bg: str):
        nonlocal row
        for c, lbl in enumerate(["Kategorie", "Monatlich", "Anteil"], 1):
            cell           = ws.cell(row=row, column=c, value=lbl)
            cell.font      = Font(name="Calibri", bold=True, size=9, color=MUTED_HEX)
            cell.fill      = _fill(bg)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left",
                                       vertical="center")
        row += 1

    def _cat_rows(cat_items, first, last, total_row, bg, color):
        nonlocal row
        cats = sorted({i.get("category_name", str(i.get("category_id", "")))
                       for i in cat_items})
        for cat in cats:
            lbl = ws.cell(row=row, column=1, value=cat)
            lbl.font   = Font(name="Calibri", size=10, color=color)
            lbl.fill   = _fill(bg)
            lbl.border = _border()

            val = ws.cell(row=row, column=2, value=_sumif_cat(first, last, cat))
            val.font          = Font(name="Calibri", size=10, bold=True, color=color)
            val.fill          = _fill(bg)
            val.border        = _border()
            val.number_format = currency_fmt
            val.alignment     = Alignment(horizontal="right", vertical="center")

            pct = ws.cell(row=row, column=3,
                          value=f"=IF(B{total_row}=0,0,B{row}/B{total_row})")
            pct.font          = Font(name="Calibri", size=10, color=color)
            pct.fill          = _fill(bg)
            pct.border        = _border()
            pct.number_format = pct_fmt
            pct.alignment     = Alignment(horizontal="right", vertical="center")
            row += 1
        row += 1

    # ── Kategorie: Einnahmen ──────────────────────────────────────────────────
    if income_items:
        _section_header("Kategorieauswertung - Einnahmen", INCOME_HEX)
        _cat_subheader(INCOME_BG)
        _cat_rows(income_items, income_first, income_last,
                  income_result_row, INCOME_BG, INCOME_HEX)

    # ── Kategorie: Ausgaben ───────────────────────────────────────────────────
    if expense_items:
        _section_header("Kategorieauswertung - Ausgaben", EXPENSE_HEX)
        _cat_subheader(EXPENSE_BG)
        _cat_rows(expense_items, expense_first, expense_last,
                  expense_result_row, EXPENSE_BG, EXPENSE_HEX)
